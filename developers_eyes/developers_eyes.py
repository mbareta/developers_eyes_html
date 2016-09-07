"""TO-DO: Write a description of what this XBlock is."""

import pkg_resources
import urllib
import os
import json

from collections import OrderedDict
from functools import partial

from xblock.core import XBlock
from xblock.fields import Scope, String
from xblock.fragment import Fragment

from openpyxl import load_workbook
from webob.response import Response
from xmodule.contentstore.content import StaticContent
from xmodule.contentstore.django import contentstore



class DevelopersEyesXBlock(XBlock):
    """
    TO-DO: document what your XBlock does.
    """

    # Fields are defined on the class.  You can access them in your code as
    # self.<fieldname>.
    display_name = String(display_name="Display Name",
                          default="Through the developers eyes interactive",
                          scope=Scope.settings,
                          help="This name appears in the horizontal navigation at the top of the page.")
    json_data = String(help="JSON data from excel file", default="None", scope=Scope.content)

    def resource_string(self, path):
        """Handy helper for getting resources from our kit."""
        data = pkg_resources.resource_string(__name__, path)
        return data.decode("utf8")

    # TO-DO: change this view to display your data your own way.
    def student_view(self, context=None):
        """
        The primary view of the DevelopersEyesXBlock, shown to students
        when viewing courses.
        """
        html = self.resource_string("static/html/developers_eyes.html")
        frag = Fragment(html.format(self=self))
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/twentytwenty.css'))
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/photo-sphere-viewer.css'))
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/developers_eyes.css'))
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/multibar_charts.css'))
        frag.add_css_url(
            self.runtime.local_resource_url(
                self, 'public/css/nvd3.css'))

        frag.add_javascript(self.resource_string("static/js/lib/zoom.js"))
        frag.add_javascript(self.resource_string("static/js/lib/jquery.event.move.js"))
        frag.add_javascript(self.resource_string("static/js/lib/jquery.twentytwenty.js"))

        # PHOTOSPHERE VIEWER
        frag.add_javascript(self.resource_string("static/js/lib/three.js"))
        frag.add_javascript(self.resource_string("static/js/lib/D.js"))
        frag.add_javascript(self.resource_string("static/js/lib/uevent.js"))
        frag.add_javascript(self.resource_string("static/js/lib/doT.js"))
        frag.add_javascript(self.resource_string("static/js/lib/photo-sphere-viewer.js"))

        # CHARTS
        frag.add_javascript(self.resource_string("static/js/lib/d3.v3.js"))
        frag.add_javascript(self.resource_string("static/js/lib/nvd3.js"))
        frag.add_javascript("var json_data ={}".format(self.json_data))

        frag.add_javascript(self.resource_string("static/js/src/developers_eyes.js"))
        frag.add_javascript(self.resource_string("static/js/src/multibar_charts.js"))

        frag.initialize_js('DevelopersEyesXBlock')
        return frag

    def studio_view(self, context):
        """
        Create a fragment used to display the edit view in the Studio.
        """
        html_str = pkg_resources.resource_string(__name__, "static/html/studio_view.html")
        frag = Fragment(unicode(html_str).format(
                                                    display_name=self.display_name,
                                                    display_description=self.display_description,
                                                    thumbnail_url=self.thumbnail_url
                                                ))
        js_str = pkg_resources.resource_string(__name__, "static/js/src/studio_edit.js")
        frag.add_javascript(unicode(js_str))
        frag.initialize_js('StudioEdit')
        return frag

    @XBlock.handler
    def studio_submit(self, request, suffix=''):
        """
        Called when submitting the form in Studio.
        """

        data = request.POST
        self.display_name = data['display_name']
        self.display_description = data['display_description']
        self.thumbnail_url = data['thumbnail_url']

        if not isinstance(data['excel'], basestring):
            upload = data['excel']

            # get workbook
            workbook = load_workbook(filename=upload.file, read_only=True)
            sheets = []
            for worksheet in workbook:
                sheet = {
                        "name": worksheet.title,
                        "rows": []
                    }
                for row in worksheet.iter_rows():
                    sheet_row = {
                        "key": None,
                        "values": []
                    }
                    cell_num = 0
                    # first row will be key for iteration, ex. "Country name"
                    for cell in row:
                        if cell_num is 0:
                            sheet_row["key"] = cell.value
                        else:
                            sheet_row["values"].append(cell.value)
                        cell_num += 1
                    sheet['rows'].append(sheet_row)

                sheets.append(sheet)

            self.json_data = json.dumps(sheets)

        return Response(json_body={
            'result': "success"
        })

    def _file_storage_name(self, filename):
        # pylint: disable=no-member
        """
        Get file path of storage.
        """
        path = (
            '{loc.block_type}/{loc.block_id}'
            '/{filename}'.format(
                    loc=self.location,
                    filename=filename
            )
        )
        return path


    # TO-DO: change this to create the scenarios you'd like to see in the
    # workbench while developing your XBlock.
    @staticmethod
    def workbench_scenarios():
        """A canned scenario for display in the workbench."""
        return [
            ("DevelopersEyesXBlock",
             """<developers_eyes/>
             """),
            ("Multiple DevelopersEyesXBlock",
             """<vertical_demo>
                <developers_eyes/>
                <developers_eyes/>
                <developers_eyes/>
                </vertical_demo>
             """),
        ]
